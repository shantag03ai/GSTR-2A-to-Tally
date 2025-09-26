import streamlit as st
import pandas as pd
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
import re

# =============================
# Helpers
# =============================

def round_amount(x):
    if x in (None, "", " ", "-") or pd.isna(x):
        return Decimal("0.00")
    try:
        return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except:
        return Decimal("0.00")

def format_date_for_tally(date_like):
    """Return YYYYMMDD (uses Note/Invoice date; never today unless truly blank/invalid)."""
    if date_like in (None, "", "-") or pd.isna(date_like):
        return datetime.now().strftime("%Y%m%d")
    if hasattr(date_like, "strftime"):
        return date_like.strftime("%Y%m%d")

    s = str(date_like).strip()
    # 2-digit years first
    for fmt in ("%d-%m-%y", "%d/%m/%y", "%d.%m.%y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year < 2000:
                dt = dt.replace(year=dt.year + 2000)
            return dt.strftime("%Y%m%d")
        except:
            pass
    # bare yyyymmdd
    if re.fullmatch(r"\d{8}", s):
        return s
    # common long formats
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m%d")
        except:
            pass
    return datetime.now().strftime("%Y%m%d")

def clean_party_name(name):
    if name in (None, "", " "):
        return "Unknown Supplier"
    s = str(name).strip()
    # remove trailing "-GSTIN"
    s = re.sub(r"\s*-\s*[A-Z0-9]{15}$", "", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s or "Unknown Supplier"

def norm_number(no):
    if no in (None, "", " "):
        return ""
    s = str(no).strip()
    s = re.sub(r"(-|\s)*(total|tot|toi)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"[^A-Za-z0-9/\-]", "", s)
    return s

def looks_like_total_or_bold(row_type_value, inv_no):
    if row_type_value in (None, "", "-", "—"):
        return True
    if re.search(r"(total|tot|toi)\s*$", str(inv_no or ""), flags=re.IGNORECASE):
        return True
    return False

# Unified schema we feed into voucher builders
B2B_COLS = [
    'GSTIN','Party_Name','Invoice_Number','Invoice_Type',
    'Invoice_Date','Invoice_Value','G','H','I','Taxable_Value',
    'IGST','CGST','SGST','CESS'
]

def _extract_rows(ws, start_row=7, take_cols=15):
    data = []
    for r in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, take_cols+1)]
        if all(v is None or str(v).strip()=="" for v in row_vals):
            continue
        data.append(row_vals)
    return data

# =============================
# Excel readers
# =============================

def extract_b2b(file_obj):
    """B2B: rows 7+, first 14 columns as per portal; map directly to schema."""
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if "B2B" not in wb.sheetnames:
            return []
        ws = wb["B2B"]
        raw = _extract_rows(ws, start_row=7, take_cols=14)
        df = pd.DataFrame(raw, columns=B2B_COLS)

        # keep valid rows
        df = df[df["GSTIN"].astype(str).str.strip().ne("")]
        df = df[df["Invoice_Number"].astype(str).str.strip().ne("")]

        seen, rows = set(), []
        for _, r in df.iterrows():
            inv_num = str(r["Invoice_Number"]).strip()
            inv_typ = str(r.get("Invoice_Type") or "").strip()
            if looks_like_total_or_bold(inv_typ, inv_num):
                continue
            key = (
                str(r["GSTIN"]).strip(),
                norm_number(inv_num),
                format_date_for_tally(r["Invoice_Date"]),
                str(round_amount(r["Taxable_Value"])),
                str(round_amount(r["IGST"])),
                str(round_amount(r["CGST"])),
                str(round_amount(r["SGST"])),
                str(round_amount(r["CESS"])),
            )
            if key in seen:
                continue
            seen.add(key)
            rows.append(r.to_dict())
        return rows
    except Exception as e:
        st.error(f"B2B read error: {e}")
        return []

def extract_cdnr(file_obj):
    """
    CDNR true portal order (based on your screenshot):
    Cols: GSTIN, Party, Note type, Note number, Note Supply type, Note date,
          Note Value, Place of supply, Reverse Charge, Rate, Taxable, IGST, CGST, SGST, Cess
    If GSTIN not present as first col in some exports, we handle that too.
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if "CDNR" not in wb.sheetnames:
            return []
        ws = wb["CDNR"]
        raw = _extract_rows(ws, start_row=7, take_cols=15)

        # Layout B (with GSTIN first) – matches your screenshot
        df_b = pd.DataFrame(raw, columns=[
            "GSTIN","Party_Name","Note_Type","Note_Number","Note_Supply_Type","Note_Date",
            "Note_Value","Place_of_supply","Reverse_Charge","Rate",
            "Taxable_Value","IGST","CGST","SGST","CESS"
        ])

        # If GSTIN column is totally blank, fall back to layout A (no GSTIN first)
        if df_b["GSTIN"].astype(str).str.strip().eq("").all():
            df_b = pd.DataFrame(raw, columns=[
                "Party_Name","Note_Type","Note_Number","Note_Supply_Type","Note_Date","Note_Value",
                "Place_of_supply","Reverse_Charge","Rate",
                "Taxable_Value","IGST","CGST","SGST","CESS","X"
            ])
            df_b["GSTIN"] = ""

        # Map to common schema (IMPORTANT: we *keep* the Note Supply Type so indexes don’t shift)
        df_b["Invoice_Number"] = df_b["Note_Number"]
        df_b["Invoice_Type"]   = df_b["Note_Type"]
        df_b["Invoice_Date"]   = df_b["Note_Date"]         # <-- correct Note date used
        df_b["Invoice_Value"]  = df_b["Note_Value"]

        df = df_b[["GSTIN","Party_Name","Invoice_Number","Invoice_Type","Invoice_Date",
                   "Invoice_Value","Rate","Place_of_supply","Reverse_Charge",
                   "Taxable_Value","IGST","CGST","SGST","CESS"]].copy()
        df.columns = B2B_COLS  # we don’t use G/H/I in vouchers; they’re placeholders

        # keep valid rows
        df = df[df["Party_Name"].astype(str).str.strip().ne("")]
        df = df[df["Invoice_Number"].astype(str).str.strip().ne("")]

        seen, rows = set(), []
        for _, r in df.iterrows():
            inv = str(r["Invoice_Number"]).strip()
            inv_typ = str(r.get("Invoice_Type") or "").strip()
            if looks_like_total_or_bold(inv_typ, inv):
                continue
            key = (
                str(r["GSTIN"]).strip(),
                norm_number(inv),
                format_date_for_tally(r["Invoice_Date"]),
                str(round_amount(r["Taxable_Value"])),
                str(round_amount(r["IGST"])),
                str(round_amount(r["CGST"])),
                str(round_amount(r["SGST"])),
                str(round_amount(r["CESS"])),
            )
            if key in seen:
                continue
            seen.add(key)
            rows.append(r.to_dict())
        return rows
    except Exception as e:
        st.error(f"CDNR read error: {e}")
        return []

# =============================
# XML utils
# =============================

def prettify(elem):
    return minidom.parseString(ET.tostring(elem, "unicode")).toprettyxml(indent="  ")

def new_env():
    env = ET.Element("ENVELOPE")
    header = ET.SubElement(env, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    return env

def add_request(env, report_name, company):
    body = ET.SubElement(env, "BODY")
    importdata = ET.SubElement(body, "IMPORTDATA")
    req = ET.SubElement(importdata, "REQUESTDESC")
    ET.SubElement(req, "REPORTNAME").text = report_name
    stat = ET.SubElement(req, "STATICVARIABLES")
    ET.SubElement(stat, "SVCURRENTCOMPANY").text = company
    return ET.SubElement(importdata, "REQUESTDATA")

def ledger_basic(tmsg_parent, name, parent, billwise="No"):
    led = ET.SubElement(tmsg_parent, "LEDGER", NAME=name, ACTION="Create")
    ET.SubElement(led, "NAME").text = name
    ET.SubElement(led, "PARENT").text = parent
    ET.SubElement(led, "ISBILLWISEON").text = billwise
    ET.SubElement(led, "OPENINGBALANCE").text = "0.00"
    ET.SubElement(led, "AFFECTSSTOCK").text = "No"
    ET.SubElement(led, "ISREVENUE").text = "Yes"
    ET.SubElement(led, "ISDEEMEDPOSITIVE").text = "No"

def ledger_tax(tmsg_parent, name, dutyhead):
    led = ET.SubElement(tmsg_parent, "LEDGER", NAME=name, ACTION="Create")
    ET.SubElement(led, "NAME").text = name
    ET.SubElement(led, "PARENT").text = "Duties & Taxes"
    ET.SubElement(led, "TAXTYPE").text = "GST"
    ET.SubElement(led, "GSTDUTYHEAD").text = dutyhead
    ET.SubElement(led, "ISINPUTCREDIT").text = "Yes"
    ET.SubElement(led, "ISBILLWISEON").text = "No"
    ET.SubElement(led, "OPENINGBALANCE").text = "0.00"

def add_bill(le, name, billtype, amount_str):
    b = ET.SubElement(le, "BILLALLOCATIONS.LIST")
    ET.SubElement(b, "NAME").text = name
    ET.SubElement(b, "BILLTYPE").text = billtype
    ET.SubElement(b, "AMOUNT").text = amount_str

# =============================
# Voucher builders (signs per Tally)
# =============================

def add_purchase_voucher(req, rec, use_inv_as_vch):
    party = clean_party_name(rec.get("Party_Name",""))
    if party == "Unknown Supplier":
        return
    inv_no = str(rec.get("Invoice_Number","")).strip()
    if not inv_no:
        return
    date = format_date_for_tally(rec.get("Invoice_Date"))

    taxable = round_amount(rec.get("Taxable_Value",0))
    igst = round_amount(rec.get("IGST",0))
    cgst = round_amount(rec.get("CGST",0))
    sgst = round_amount(rec.get("SGST",0))
    cess = round_amount(rec.get("CESS",0))
    total = taxable + igst + cgst + sgst + cess
    if total <= 0:
        return

    tmsg = ET.SubElement(req, "TALLYMESSAGE", {"xmlns:UDF":"TallyUDF"})
    v = ET.SubElement(tmsg, "VOUCHER", VCHTYPE="Purchase", ACTION="Create")
    ET.SubElement(v, "VOUCHERTYPENAME").text = "Purchase"
    ET.SubElement(v, "DATE").text = date
    ET.SubElement(v, "EFFECTIVEDATE").text = date
    ET.SubElement(v, "VOUCHERNUMBER").text = inv_no if (use_inv_as_vch and inv_no) else ""
    ET.SubElement(v, "REFERENCE").text = inv_no
    ET.SubElement(v, "NARRATION").text = f"Purchase from {party} - Invoice: {inv_no}"
    ET.SubElement(v, "PARTYLEDGERNAME").text = party

    # DR lines (negative)
    if taxable > 0:
        le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(le, "LEDGERNAME").text = "Purchase Taxable"
        ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "Yes"
        ET.SubElement(le, "AMOUNT").text = f"-{taxable}"
        add_bill(le, inv_no, "New Ref", f"-{taxable}")

    for ln, amt in [("INPUT IGST", igst), ("INPUT CGST", cgst), ("INPUT SGST", sgst), ("INPUT CESS", cess)]:
        if amt > 0:
            le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
            ET.SubElement(le, "LEDGERNAME").text = ln
            ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "Yes"
            ET.SubElement(le, "AMOUNT").text = f"-{amt}"

    # CR party (positive)
    le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(le, "LEDGERNAME").text = party
    ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "No"
    ET.SubElement(le, "AMOUNT").text = f"{total}"
    add_bill(le, inv_no, "New Ref", f"{total}")

def add_cdnr_voucher(req, rec, use_inv_as_vch):
    party = clean_party_name(rec.get("Party_Name",""))
    if party == "Unknown Supplier":
        return
    note_no = str(rec.get("Invoice_Number","")).strip()
    if not note_no:
        return
    date = format_date_for_tally(rec.get("Invoice_Date"))  # <-- uses NOTE DATE

    taxable = round_amount(rec.get("Taxable_Value",0))
    igst = round_amount(rec.get("IGST",0))
    cgst = round_amount(rec.get("CGST",0))
    sgst = round_amount(rec.get("SGST",0))
    cess = round_amount(rec.get("CESS",0))
    total = taxable + igst + cgst + sgst + cess
    if total <= 0:
        return

    note_type = str(rec.get("Invoice_Type","")).strip().lower()
    # Supplier Credit Note -> our Debit Note (reverse purchase)
    vch_type = "Debit Note"
    narration = "Debit Note (Purchase Return)"
    if "debit" in note_type:
        vch_type = "Credit Note"
        narration = "Credit Note (Supplier Debit)"

    tmsg = ET.SubElement(req, "TALLYMESSAGE", {"xmlns:UDF":"TallyUDF"})
    v = ET.SubElement(tmsg, "VOUCHER", VCHTYPE=vch_type, ACTION="Create")
    ET.SubElement(v, "VOUCHERTYPENAME").text = vch_type
    ET.SubElement(v, "VCHENTRYMODE").text = "Accounting"
    ET.SubElement(v, "DATE").text = date
    ET.SubElement(v, "EFFECTIVEDATE").text = date
    ET.SubElement(v, "VOUCHERNUMBER").text = note_no if (use_inv_as_vch and note_no) else ""
    ET.SubElement(v, "REFERENCE").text = note_no
    ET.SubElement(v, "NARRATION").text = f"{narration} - Supplier: {party} - Ref: {note_no}"
    ET.SubElement(v, "PARTYLEDGERNAME").text = party

    if vch_type == "Debit Note":
        # Party Dr (negative), Purchase/Taxes Cr (positive)
        le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(le, "LEDGERNAME").text = party
        ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "Yes"
        ET.SubElement(le, "AMOUNT").text = f"-{total}"
        add_bill(le, note_no, "Agst Ref", f"-{total}")

        if taxable > 0:
            le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
            ET.SubElement(le, "LEDGERNAME").text = "Purchase Taxable"
            ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "No"
            ET.SubElement(le, "AMOUNT").text = f"{taxable}"
        for ln, amt in [("INPUT IGST", igst), ("INPUT CGST", cgst), ("INPUT SGST", sgst), ("INPUT CESS", cess)]:
            if amt > 0:
                le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
                ET.SubElement(le, "LEDGERNAME").text = ln
                ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "No"
                ET.SubElement(le, "AMOUNT").text = f"{amt}"
    else:
        # Supplier Debit Note -> our Credit Note (Party Cr; Purchase/Taxes Dr)
        le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(le, "LEDGERNAME").text = party
        ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(le, "AMOUNT").text = f"{total}"
        add_bill(le, note_no, "Agst Ref", f"{total}")

        if taxable > 0:
            le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
            ET.SubElement(le, "LEDGERNAME").text = "Purchase Taxable"
            ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "Yes"
            ET.SubElement(le, "AMOUNT").text = f"-{taxable}"
        for ln, amt in [("INPUT IGST", igst), ("INPUT CGST", cgst), ("INPUT SGST", sgst), ("INPUT CESS", cess)]:
            if amt > 0:
                le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
                ET.SubElement(le, "LEDGERNAME").text = ln
                ET.SubElement(le, "ISDEEMEDPOSITIVE").text = "Yes"
                ET.SubElement(le, "AMOUNT").text = f"-{amt}"

# =============================
# Build ONE combined XML (masters first, then all vouchers)
# =============================

def build_one_combined_xml(company, b2b_rows, cdnr_rows, use_inv_as_vch):
    env = new_env()

    # ---- Masters block
    req_m = add_request(env, "All Masters", company)
    tm_m = ET.SubElement(req_m, "TALLYMESSAGE", {"xmlns:UDF":"TallyUDF"})
    # Standard ledgers
    ledger_basic(tm_m, "Purchase Taxable", "Purchase Accounts")
    ledger_basic(tm_m, "Purchase Nil Rated", "Purchase Accounts")
    ledger_tax(tm_m, "INPUT IGST", "Integrated Tax")
    ledger_tax(tm_m, "INPUT CGST", "Central Tax")
    ledger_tax(tm_m, "INPUT SGST", "State Tax")
    ledger_tax(tm_m, "INPUT CESS", "Cess")
    ledger_basic(tm_m, "Rounding Off", "Indirect Expenses")
    # Supplier ledgers (billwise Yes)
    suppliers = {clean_party_name(r.get("Party_Name","")) for r in (b2b_rows + cdnr_rows)}
    for s in sorted(suppliers):
        if s and s != "Unknown Supplier":
            ledger_basic(tm_m, s, "Sundry Creditors", billwise="Yes")

    # ---- Vouchers block
    req_v = add_request(env, "Vouchers", company)
    for rec in b2b_rows:
        add_purchase_voucher(req_v, rec, use_inv_as_vch)
    for rec in cdnr_rows:
        add_cdnr_voucher(req_v, rec, use_inv_as_vch)

    return prettify(env)

# =============================
# Streamlit App (ONE download)
# =============================

def main():
    st.set_page_config(page_title="GSTR-2A → Tally XML (One File)", page_icon="📊", layout="wide")
    st.markdown("""
    <div style="background: linear-gradient(90deg, #FFD700, #FFF8DC); padding: 15px; border-radius: 10px; margin-bottom: 30px; text-align: center;">
      <h2 style="color: #333; margin: 0; font-weight: bold;">📊 Developed by CA Shantam Jagdish Agrawal 📊</h2>
    </div>
    """, unsafe_allow_html=True)

    st.title("🔄 GSTR-2A EXCEL → Tally XML (Purchase & Debit/Credit Notes)")

    company = st.text_input("🏢 Tally Company Name (exact)", "")
    use_inv_as_vch = st.checkbox("Use Invoice/Note No as Voucher No", value=True)
    files = st.file_uploader("Upload one or more GSTR-2A Excel files (portal export)", type=["xlsx","xls"], accept_multiple_files=True)
    out_name = st.text_input("Output XML filename", "gstr2a_all_in_one.xml")

    if st.button("🚀 Generate Single XML", type="primary", use_container_width=True):
        if not company or not files:
            st.error("Enter company name and upload at least one file.")
            return

        all_b2b, all_cdnr = [], []
        for f in files:
            f.seek(0); all_b2b.extend(extract_b2b(f))
            f.seek(0); all_cdnr.extend(extract_cdnr(f))

        st.success(f"✅ Parsed — B2B rows: {len(all_b2b)} | CDNR rows: {len(all_cdnr)}")

        combined_xml = build_one_combined_xml(company, all_b2b, all_cdnr, use_inv_as_vch)

        st.download_button("💾 Download Combined XML (Masters + Vouchers)",
                           data=combined_xml, file_name=out_name, mime="text/xml",
                           use_container_width=True)

        st.info("""**Import in Tally Prime**
- Import this **single XML**. It creates required ledgers first, then posts **Purchase** (B2B) and **Debit/Credit Notes** (CDNR).
- CDNR date uses the **Note date** from your sheet.
- For supplier *Credit Note* → **Debit Note** (Party Dr; Purchase/Taxes Cr).  
  For supplier *Debit Note* → **Credit Note** (Party Cr; Purchase/Taxes Dr).""")

if __name__ == "__main__":
    main()
